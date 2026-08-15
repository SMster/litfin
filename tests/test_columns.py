"""Stage, court, counsel, the settlement rebalance, and the Excel export.

The recurring theme: a table cell is read without its caveats. The dashboard
can put a banner above an imputed figure; a spreadsheet cell that gets sorted,
summed and mailed onward cannot. So each of these tests is really asking the
same question — does this column stay honest once it leaves the screen?
"""

from __future__ import annotations

from pathlib import Path

import pytest

from litfin.deliver import dataset, excel
from litfin.deliver.dataset import STAGE_ORDER, STAGE_UNKNOWN, derive_stage
from litfin.score.scoring import EVENT_FIT, THESIS_PRIORITY, score_row


# ---------------------------------------------------------------------------
# Scoring: near/post settlement outranks judgment
# ---------------------------------------------------------------------------

def _row(**kw):
    base = dict(
        deal_thesis="post_settlement", event_type="settlement_reached",
        practice_area="commercial", venue="S.D.N.Y.", court="", jurisdiction="federal",
        event_date="2026-08-01", published_at="2026-08-01",
        damages_usd=None, damages_conf="none", payload_json="{}",
    )
    base.update(kw)
    return base


class TestSettlementWeighting:
    def test_approved_settlement_outranks_an_entered_judgment(self):
        """A settled case has an agreed number and a payer who has already
        decided not to fight. A judgment still faces post-trial motions,
        appeal, and collection."""
        assert EVENT_FIT["settlement_final_approval"] > EVENT_FIT["judgment_entered"]
        assert EVENT_FIT["settlement_reached"] > EVENT_FIT["judgment_entered"]

    def test_a_verdict_is_not_treated_as_de_risked(self):
        """The most appealable moment in a case's life."""
        assert EVENT_FIT["jury_verdict"] < EVENT_FIT["judgment_entered"]
        assert EVENT_FIT["jury_verdict"] < EVENT_FIT["settlement_reached"]

    def test_settlement_ordering_runs_final_then_prelim_then_agreed(self):
        assert (
            EVENT_FIT["settlement_final_approval"]
            > EVENT_FIT["settlement_preliminary_approval"]
            > EVENT_FIT["settlement_reached"]
        )

    def test_post_settlement_thesis_is_prioritized(self):
        assert THESIS_PRIORITY["post_settlement"] == 1.0
        assert THESIS_PRIORITY["antitrust_followon"] < 1.0
        assert THESIS_PRIORITY["none"] == 0.0

    def test_end_to_end_a_settlement_beats_a_judgment(self):
        settled, _ = score_row(_row(
            deal_thesis="post_settlement", event_type="settlement_final_approval"))
        judged, _ = score_row(_row(
            deal_thesis="judgment_monetization", event_type="judgment_entered"))
        assert settled > judged

    def test_off_thesis_still_scores_zero_thesis_fit(self):
        _total, c = score_row(_row(deal_thesis="none",
                                   event_type="settlement_final_approval"))
        assert c.thesis_fit == 0.0

    def test_antitrust_followon_is_not_buried(self):
        """The rebalance must not cost the antitrust thesis its place -- it is
        one of the three, and its events are usually judgment_proposed."""
        anti, _ = score_row(_row(deal_thesis="antitrust_followon",
                                 event_type="settlement_reached"))
        post, _ = score_row(_row(deal_thesis="post_settlement",
                                 event_type="settlement_reached"))
        assert anti == pytest.approx(post, rel=0.08)

    def test_event_fit_is_overridable_per_call(self):
        """litfin.toml [score.event_fit] has to actually reach the scorer."""
        low, _ = score_row(_row(event_type="settlement_reached"),
                           event_fit={"settlement_reached": 0.1})
        high, _ = score_row(_row(event_type="settlement_reached"))
        assert low < high


class TestConfigWiring:
    def test_toml_weights_reach_the_config(self, tmp_path):
        """The scoring docstring promised these were tunable in litfin.toml
        from the start, and nothing ever loaded them."""
        from litfin.config import load_config

        cfg_file = tmp_path / "litfin.toml"
        cfg_file.write_text(
            'purpose = "research"\n'
            "[score.weights]\ndamages = 0.5\n"
            "[score.event_fit]\nsettlement_reached = 0.42\n",
            encoding="utf-8",
        )
        cfg = load_config(cfg_file)
        assert cfg.weights == {"damages": 0.5}
        assert cfg.event_fit == {"settlement_reached": 0.42}

    def test_absent_score_section_means_defaults(self, tmp_path):
        from litfin.config import load_config

        cfg_file = tmp_path / "litfin.toml"
        cfg_file.write_text('purpose = "research"\n', encoding="utf-8")
        cfg = load_config(cfg_file)
        assert cfg.weights == {} and cfg.event_fit == {}


# ---------------------------------------------------------------------------
# Stage
# ---------------------------------------------------------------------------

class TestStage:
    def test_posture_beats_event_type(self):
        """event_type says what JUST HAPPENED; procedural_posture says where
        the case STANDS, and the second is the question a triager asks."""
        stage, basis = derive_stage(
            "Motion to dismiss is fully briefed and pending.", "case_filed"
        )
        assert stage == "Motion to dismiss"
        assert basis == "posture"

    def test_event_type_is_the_fallback_and_says_so(self):
        stage, basis = derive_stage("", "settlement_final_approval")
        assert stage == "Settlement — final approval"
        assert basis == "event type"

    def test_unknown_when_neither_is_specific(self):
        stage, basis = derive_stage("", "")
        assert stage == STAGE_UNKNOWN
        assert "neither" in basis

    @pytest.mark.parametrize("posture,expected", [
        ("Defendant's motion to dismiss under Rule 12(b)(6) is pending",
         "Motion to dismiss"),
        ("The parties are engaged in discovery; depositions are ongoing",
         "Discovery"),
        ("Cross-motions for summary judgment are pending", "Summary judgment"),
        ("Trial is scheduled for October", "Trial"),
        ("The jury returned a verdict for plaintiff", "Verdict returned"),
        ("Post-trial motions are pending", "Post-trial motions"),
        ("Final judgment was entered against the defendant", "Judgment entered"),
        ("Notice of appeal filed to the Ninth Circuit", "On appeal"),
        ("The parties reached a settlement in principle", "Settlement reached"),
        ("Settlement is subject to preliminary approval", "Settlement — preliminary approval"),
        ("The court granted final approval of the settlement",
         "Settlement — final approval"),
        ("Referred to mediation", "Settlement talks"),
        ("Chapter 11 plan confirmed", "Plan confirmation"),
        ("Case closed", "Closed"),
    ])
    def test_posture_patterns(self, posture, expected):
        assert derive_stage(posture, "no_event")[0] == expected

    def test_latest_stage_wins_within_one_posture(self):
        """A posture mentioning both a settlement and an earlier motion is at
        the settlement stage. Scanning latest-first gets that right without
        parsing the sentence."""
        stage, _ = derive_stage(
            "After the motion to dismiss was denied and discovery closed, the "
            "parties reached a settlement.",
            "no_event",
        )
        assert stage == "Settlement reached"

    def test_every_derived_stage_is_in_the_order_list(self):
        """A stage missing from STAGE_ORDER sorts to the end silently."""
        from litfin.deliver.dataset import _STAGE_FROM_EVENT, _STAGE_PATTERNS

        for label, _ in _STAGE_PATTERNS:
            assert label in STAGE_ORDER, label
        for label in _STAGE_FROM_EVENT.values():
            assert label in STAGE_ORDER, label

    def test_order_is_by_fundability_not_chronology(self):
        """Sorting the column descending should surface the most fundable
        matters. 'Closed' is chronologically last and commercially least
        interesting, so it must NOT sort to the top."""
        rank = dataset.stage_rank
        assert rank("Settlement — final approval") > rank("Judgment entered")
        assert rank("Judgment entered") > rank("On appeal")
        assert rank("Closed") < rank("Pleadings")
        assert rank(STAGE_UNKNOWN) < rank("Closed")

    def test_unknown_stage_rank_is_stable(self):
        assert dataset.stage_rank("something invented") == len(STAGE_ORDER)


# ---------------------------------------------------------------------------
# Counsel
# ---------------------------------------------------------------------------

class TestCounsel:
    def test_schema_carries_counsel_and_says_not_to_guess(self):
        from litfin.extract.schema import CaseExtraction

        fields = CaseExtraction.model_fields
        assert "counsel_plaintiff" in fields and "counsel_defendant" in fields
        for name in ("counsel_plaintiff", "counsel_defendant"):
            desc = fields[name].description or ""
            assert "do NOT infer or guess" in desc
            assert "FIRMS" in desc

    def test_schema_version_bumped_so_refresh_can_find_old_rows(self):
        from litfin.extract.schema import SCHEMA_VERSION

        assert SCHEMA_VERSION >= 2

    def test_prompt_tells_the_model_empty_is_the_usual_answer(self):
        from litfin.extract.prompts import SYSTEM_PROMPT

        assert "Counsel: firms only, and usually empty" in SYSTEM_PROMPT
        assert "An empty list is the" in SYSTEM_PROMPT

    def test_not_named_and_not_captured_are_different_cells(self):
        """Only one of the two is fixable by re-extracting, so a reader must
        be able to tell them apart."""
        assert excel._counsel([], known=True) == "not named in document"
        assert "pre-v2" in excel._counsel([], known=False)
        assert excel._counsel(["Quinn Emanuel"], known=True) == "Quinn Emanuel"

    def test_dashboard_distinguishes_them_too(self):
        from litfin.deliver import dashboard

        js = dashboard._JS
        assert "not named" in js and "not captured" in js
        assert "p.counsel_known" in js


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _dataset_with(**kw):
    from test_deliver import make_dataset, make_prospect

    return make_dataset(prospects=[make_prospect(**kw)])


class TestExcelExport:
    def _load(self, path):
        from openpyxl import load_workbook

        return load_workbook(path)

    def test_writes_three_sheets(self, tmp_path):
        from test_deliver import make_dataset

        out = excel.build(make_dataset(), tmp_path / "x.xlsx")
        wb = self._load(out)
        assert wb.sheetnames == ["Prospects", "Venue coverage", "Sources"]

    def test_stated_damages_land_in_a_numeric_cell(self, tmp_path):
        out = excel.build(
            _dataset_with(damages_usd=40_000_000.0, damages_imputed=False),
            tmp_path / "x.xlsx",
        )
        ws = self._load(out)["Prospects"]
        assert ws.cell(row=7, column=12).value == 40_000_000.0
        assert ws.cell(row=7, column=12).number_format == excel.MONEY_FORMAT

    def test_imputed_damages_are_NEVER_written_to_the_numeric_column(self, tmp_path):
        """THE important one. A spreadsheet gets summed, averaged and charted
        away from every caveat the dashboard put around the number."""
        out = excel.build(
            _dataset_with(damages_usd=15_000_000.0, damages_imputed=True),
            tmp_path / "x.xlsx",
        )
        ws = self._load(out)["Prospects"]
        assert ws.cell(row=7, column=12).value is None
        assert ws.cell(row=7, column=13).value == dataset.NOT_STATED

    def test_the_header_block_explains_the_blank(self, tmp_path):
        from test_deliver import make_dataset

        ws = self._load(
            excel.build(make_dataset(), tmp_path / "x.xlsx")
        )["Prospects"]
        assert "STATED figures only" in ws["A3"].value

    def test_coverage_warning_travels_into_the_file(self, tmp_path):
        """An exported file travels away from the banner that would have
        explained an empty venue."""
        from test_deliver import make_dataset

        wb = self._load(excel.build(make_dataset(), tmp_path / "x.xlsx"))
        assert "no PACER RSS feed" in wb["Prospects"]["A4"].value
        cov = wb["Venue coverage"]
        assert "trust an EMPTY result" in cov["A1"].value
        assert any(
            (cov.cell(row=r, column=1).value == "low") for r in range(5, 10)
        )

    def test_broken_sources_are_visible_in_the_export(self, tmp_path):
        from test_deliver import make_dataset
        from litfin.deliver.dataset import SourceRow

        data = make_dataset(sources=[
            SourceRow("sec_fts", "SEC", "A", "gov", "BROKEN", "note", "", 3, 0),
        ])
        ws = self._load(excel.build(data, tmp_path / "x.xlsx"))["Sources"]
        assert ws.cell(row=5, column=4).value == "BROKEN"

    def test_header_is_frozen_and_filterable(self, tmp_path):
        from test_deliver import make_dataset

        ws = self._load(
            excel.build(make_dataset(), tmp_path / "x.xlsx")
        )["Prospects"]
        assert ws.freeze_panes == "D7"
        assert ws.auto_filter.ref is not None

    def test_inferred_stage_is_marked(self, tmp_path):
        from test_deliver import make_dataset, make_prospect

        p = make_prospect()
        p.stage = "Judgment entered"
        p.stage_basis = "event type"
        ws = self._load(
            excel.build(make_dataset(prospects=[p]), tmp_path / "x.xlsx")
        )["Prospects"]
        assert "(inferred)" in ws.cell(row=7, column=6).value

    def test_url_becomes_a_hyperlink(self, tmp_path):
        from test_deliver import make_dataset

        ws = self._load(
            excel.build(make_dataset(), tmp_path / "x.xlsx")
        )["Prospects"]
        assert ws.cell(row=7, column=22).hyperlink is not None

    def test_empty_dataset_still_produces_a_valid_file(self, tmp_path):
        from test_deliver import make_dataset

        out = excel.build(make_dataset(prospects=[]), tmp_path / "x.xlsx")
        assert out.is_file()
        assert self._load(out)["Prospects"].auto_filter.ref is None

    def test_column_headers_match_the_values_written(self, tmp_path):
        """A mismatch here silently shifts every value one column left."""
        from test_deliver import make_dataset

        ws = self._load(
            excel.build(make_dataset(), tmp_path / "x.xlsx")
        )["Prospects"]
        headers = [
            ws.cell(row=6, column=i + 1).value for i in range(len(excel.COLUMNS))
        ]
        assert headers == [c[0] for c in excel.COLUMNS]
        assert ws.cell(row=6, column=len(excel.COLUMNS) + 1).value is None


class TestServerExportRoute:
    def test_route_exists_and_sets_a_download_filename(self):
        from litfin.deliver import server

        src = Path(server.__file__).read_text(encoding="utf-8")
        assert "/export.xlsx" in src
        assert "attachment; filename=" in src
        assert "spreadsheetml.sheet" in src

    def test_export_is_a_free_action_needing_no_confirmation(self):
        from litfin.config import Config
        from litfin.deliver import server

        panel = server._panel_html(Config(), "tok")
        assert 'href="/export.xlsx"' in panel
        assert "data-confirm" not in panel.split("export.xlsx")[0].rsplit("<a", 1)[-1]
