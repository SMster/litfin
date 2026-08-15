"""Excel export of the ranked table.

Fed by the same `dataset.load()` as the dashboard and the digest, so the
workbook cannot disagree with the screen it was exported from.

THE ONE RULE THAT MATTERS: a spreadsheet gets filtered, sorted, pivoted and
mailed onward, detached from every caveat the dashboard puts around a number.
So the distinctions the UI makes carefully must survive into the cells, not be
flattened into a plausible-looking blank:

  * an IMPUTED damages figure is never written into the numeric Damages
    column -- that column holds stated figures only, and the imputed value
    would otherwise be summed, averaged and charted as if it were real
  * "no counsel named in the document" and "this row predates counsel
    capture" are different cells, because only one is fixable
  * a stage inferred from an event type is marked, since it is a weaker claim
    than one read off an explicit procedural posture

A second sheet carries the venue coverage map, because "no results in D. Nev."
means nothing without it and an exported file travels away from the banner
that would have said so.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from .dataset import Dataset

# openpyxl is imported lazily inside build() so that importing this module --
# which `deliver/__init__` does -- never hard-fails an install that has no
# Excel support.
_HEADER_FILL = "FF1A4F8A"
_ZEBRA_FILL = "FFF4F6F9"
_WARN_FONT = "FF8A5A00"
_MUTED_FONT = "FF6B7280"

# (header, width, wrap). Order is the column order in the sheet.
COLUMNS: tuple[tuple[str, int, bool], ...] = (
    ("Rank", 6, False),
    ("Score", 8, False),
    ("Case", 42, True),
    ("What happened", 52, True),
    ("Summary", 64, True),
    ("Stage", 26, False),
    ("Court", 34, True),
    ("Venue", 16, False),
    ("Jurisdiction", 20, False),
    ("Plaintiff counsel", 30, True),
    ("Defendant counsel", 30, True),
    ("Damages (stated)", 18, False),
    ("Claim size band", 20, False),
    ("Damages basis", 34, True),
    ("Thesis", 22, False),
    ("Event type", 26, False),
    ("Event date", 13, False),
    ("Defendants", 30, True),
    ("Public defendant", 15, False),
    ("Documents", 11, False),
    ("Source", 20, False),
    ("URL", 46, False),
)

MONEY_FORMAT = '"$"#,##0'


def _counsel(firms: list[str], known: bool) -> str:
    if firms:
        return "; ".join(firms)
    # Two different facts, two different cells. Collapsing them into "" would
    # make an unfixable gap look identical to a fixable one.
    return "not named in document" if known else "not captured (pre-v2 extraction)"


def build(data: Dataset, out_path: Path, *, limit: int | None = None) -> Path:
    """Write the ranked table to a formatted .xlsx. Returns the path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = data.prospects[:limit] if limit else data.prospects

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    zebra = PatternFill("solid", fgColor=_ZEBRA_FILL)
    warn_font = Font(color=_WARN_FONT, size=10, italic=True)
    muted_font = Font(color=_MUTED_FONT, size=10, italic=True)
    top_wrap = Alignment(vertical="top", wrap_text=True)
    top_plain = Alignment(vertical="top")

    # -- title block, so an exported file explains itself ------------------
    ws["A1"] = f"LitFin prospects — {data.generated_at[:10]}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"{len(rows)} matters · declared purpose: {data.purpose} · "
        f"generated {data.generated_at}"
    )
    ws["A2"].font = muted_font
    ws["A3"] = (
        "Damages holds STATED figures only. A blank means no figure appeared "
        "in the source — see 'Claim size band' = 'Not stated'. Rankings for "
        "those rows rest on a thesis prior, not a number, and the value must "
        "not be treated as an estimate."
    )
    ws["A3"].font = warn_font
    if data.dark_venues or data.partial_venues:
        ws["A4"] = (
            f"Venue coverage is incomplete: {data.dark_venues} courts publish "
            f"no PACER RSS feed and {data.partial_venues} publish "
            f"orders/opinions only. In those venues an empty result is "
            f"absence of signal, not absence of activity — see the "
            f"'Venue coverage' sheet."
        )
        ws["A4"].font = warn_font

    header_row = 6
    for col, (title, width, _wrap) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[header_row].height = 26

    for i, p in enumerate(rows):
        r = header_row + 1 + i
        event_date = (p.event_date or p.published_at or "")[:10]
        try:
            event_val: Any = date.fromisoformat(event_date) if event_date else ""
        except ValueError:
            event_val = event_date

        values = [
            p.rank,
            round(p.score, 4),
            p.caption or "(untitled)",
            p.description,
            p.summary,
            p.stage + ("" if p.stage_basis == "posture" else "  (inferred)"),
            p.court_display,
            p.venue,
            p.jurisdiction_label,
            _counsel(p.counsel_plaintiff, p.counsel_known),
            _counsel(p.counsel_defendant, p.counsel_known),
            # THE IMPORTANT ONE: never write an imputed figure into a numeric
            # column. It would be summed and charted as if it were real.
            (p.damages_usd if (p.damages_usd and not p.damages_imputed) else None),
            p.size_band,
            p.damages_basis,
            p.deal_thesis,
            p.event_type,
            event_val,
            "; ".join(p.parties_defendant),
            "yes" if p.defendant_is_public else "",
            p.document_count,
            p.source_id,
            p.source_url,
        ]

        for col, (value, (_title, _w, wrap)) in enumerate(
            zip(values, COLUMNS), start=1
        ):
            cell = ws.cell(row=r, column=col, value=value)
            cell.alignment = top_wrap if wrap else top_plain
            if i % 2:
                cell.fill = zebra

        ws.cell(row=r, column=12).number_format = MONEY_FORMAT
        if event_val and not isinstance(event_val, str):
            ws.cell(row=r, column=17).number_format = "yyyy-mm-dd"
        if p.damages_imputed:
            ws.cell(row=r, column=13).font = warn_font
        if p.stage_basis != "posture":
            ws.cell(row=r, column=6).font = muted_font
        if not (p.counsel_plaintiff or p.counsel_defendant):
            ws.cell(row=r, column=10).font = muted_font
            ws.cell(row=r, column=11).font = muted_font
        url = ws.cell(row=r, column=22)
        if p.source_url:
            url.hyperlink = p.source_url
            url.font = Font(color="FF1A4F8A", underline="single", size=10)

    last_row = header_row + len(rows)
    if rows:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(COLUMNS))}{last_row}"
        )
    # Freeze the header AND the rank/score/case columns, so scrolling right to
    # the counsel columns does not lose track of which case you are reading.
    ws.freeze_panes = f"D{header_row + 1}"

    _coverage_sheet(wb, data)
    _sources_sheet(wb, data)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _coverage_sheet(wb, data: Dataset) -> None:
    """The venue coverage map. An exported file travels away from the banner
    that would have explained an empty venue."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Venue coverage")
    ws["A1"] = "How much to trust an EMPTY result, per court"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "A court with no PACER RSS feed produces no rows whether or not "
        "anything happened in it."
    )
    ws["A2"].font = Font(color=_WARN_FONT, italic=True, size=10)

    headers = ("Confidence", "Court ID", "Court", "Jurisdiction", "Entry types")
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=title)
        c.font = Font(bold=True, color="FFFFFFFF", size=10)
        c.fill = fill
    for width, col in zip((14, 12, 46, 14, 26), "ABCDE"):
        ws.column_dimensions[col].width = width

    for i, court in enumerate(data.courts):
        r = 5 + i
        ws.cell(row=r, column=1, value=court.confidence)
        ws.cell(row=r, column=2, value=court.court_id)
        ws.cell(row=r, column=3, value=court.full_name).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=r, column=4, value=court.jurisdiction)
        ws.cell(row=r, column=5, value=court.entry_types)
        if court.confidence == "low":
            ws.cell(row=r, column=1).font = Font(color="FF98221F", bold=True)
    if data.courts:
        ws.auto_filter.ref = f"A4:E{4 + len(data.courts)}"
    ws.freeze_panes = "A5"


def _sources_sheet(wb, data: Dataset) -> None:
    """Source health, so a BROKEN source is visible in the export too."""
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Sources")
    ws["A1"] = "Source health at export time"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "A source that is not HEALTHY may be missing entirely from the "
        "Prospects sheet. That is not a quiet day."
    )
    ws["A2"].font = Font(color=_WARN_FONT, italic=True, size=10)

    headers = ("Source", "Tier", "ToS status", "Health", "Items",
               "Last success", "Note")
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=title)
        c.font = Font(bold=True, color="FFFFFFFF", size=10)
        c.fill = fill
    for width, col in zip((24, 6, 20, 12, 9, 26, 50), "ABCDEFG"):
        ws.column_dimensions[col].width = width

    for i, s in enumerate(data.sources):
        r = 5 + i
        ws.cell(row=r, column=1, value=s.source_id)
        ws.cell(row=r, column=2, value=s.tier)
        ws.cell(row=r, column=3, value=s.status)
        health = ws.cell(row=r, column=4, value=s.health)
        if s.health != "HEALTHY":
            health.font = Font(color="FF98221F", bold=True)
        ws.cell(row=r, column=5, value=s.items)
        ws.cell(row=r, column=6, value=s.last_success_at[:19])
        ws.cell(row=r, column=7, value=s.health_note[:200])
    ws.freeze_panes = "A5"


def default_path(cfg, stamp: str | None = None) -> Path:
    day = stamp or datetime.now().strftime("%Y-%m-%d")
    return cfg.data_root / f"litfin-prospects-{day}.xlsx"
